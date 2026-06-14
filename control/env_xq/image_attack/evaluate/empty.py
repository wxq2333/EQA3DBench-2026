import openpyxl

def save_text_from_empty_next_row(file_path, specified_rows, output_txt_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    saved_texts = []
    for row in specified_rows:  # 遍历指定的多行
        for j in range(1, sheet.max_column):  # 从第1列开始遍历到最大列数
            current_cell = sheet.cell(row=row, column=j)
            next_row_cell = sheet.cell(row=row + 1, column=j)
            if next_row_cell.value is None:  # 检查下一行对应列的单元格是否为空
                saved_texts.append(current_cell.value)
    # 将结果保存到指定的 txt 文件中
    with open(output_txt_path, 'w') as f:
        for text in saved_texts:
            if text is not None:  # 避免写入 None 值
                f.write('"' + str(text) + '"\n')


# 示例调用

# 示例调用
file_path = '/home/ubuntu/下载/副本数据集2.9.xlsx'
specified_rows = [42]  # 假设你指定的行是第 3、5、7 行
output_txt_path = '/home/ubuntu/桌面/control/env_xq/image_attack/evaluate/empty.txt'  # 指定输出的 txt 文件路径
save_text_from_empty_next_row(file_path, specified_rows, output_txt_path)